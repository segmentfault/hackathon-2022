from tqdm import tqdm
from openpyxl import Workbook
from openpyxl import load_workbook
from collections import defaultdict
import jieba
import time
from  boson_dict_config import bonson_config as bc
import matplotlib.pyplot as plt

def load_jieba(compile_sentiment_dict):
    """
    add compile_sentiment_dict into jieba
    :param compile_sentiment_dict: {'compile phrase':score}
    :return: add compile words into jieba dictionary
    """
    for words in compile_sentiment_dict:
        jieba.add_word(words)

def get_sentiment_dict(boson_dict, sentiment_dir1, sentiment_dir2, sentiment_dir3, sentiment_dir4):
    """
    get sentiment dict and compile dict
    :param boson_dict: {'Words':'Boson_Score'}
    :return: {'compile phrase':score},{'pos':[...],'neg':[...]}
    """
    sentiment_dict = {}
    re_sentiment_dict = {}
    # print(boson_dict.keys())
    file = open(sentiment_dir1, 'r', encoding='utf-8').read()
    file2 = open(sentiment_dir2, 'r', encoding='utf-8').read()
    file3 = open(sentiment_dir3, 'r', encoding='utf-8').read()
    file4 = open(sentiment_dir4, 'r', encoding='utf-8').read()
    sentiment_dict['pos'] = [word.strip() for word in file.split('\n')[1::] if word != '']
    sentiment_dict['neg'] = [word.strip() for word in file3.split('\n')[1::] if word != '']
    for word in file2.split('\n')[1::]:
        if word != '':
            sentiment_dict['pos'].append(word.strip())
    for word in file4.split('\n')[1::]:
        if word != '':
            sentiment_dict['neg'].append(word.strip())
    # print ("=========>", sentiment_dict.keys())
    # print ("sentiment_dict.keys() ",sentiment_dict.keys())
    for flag in sentiment_dict:
        for phrase in sentiment_dict[flag]:
            if phrase in boson_dict.keys():
                re_sentiment_dict[phrase] = boson_dict[phrase]
    # print('result:',re_sentiment_dict)
    return re_sentiment_dict,sentiment_dict

def get_adverb_dict():
    """
    get adverb_dict
    :return: {'extreme':[...],...,'insufficently':[...]}
    """
    adverb_dict = {}
    adverb_dict['extreme'] = ['????????????', '??????', '??????', '?????????', '??????', '????????????', '????????????', '????????????',
                            '????????????', '??????', '??????', '????????????', '??????', '???', '??????', '??????', '??????', '??????',
                            '??????', '???', '?????????', '???', '??????', '??????', '?????????', '??????', '???', '???', '??????',
                            '??????', '??????', '???', '??????', '??????', '?????????', '??????', '??????', '???', '??????', '???',
                            '???', '??????', '????????????', '???', '??????', '??????', '??????', '??????', '??????', '????????????',
                            '????????????', '????????????', '??????', '??????', '??????', '??????', '??????', '??????', '???', '??????',
                            '??????', '??????', '??????', '??????', '??????', '???', '???', '???', '???']
    adverb_dict['very'] = ['??????', '??????', '??????', '???', '???', '??????', '??????', '??????', '???', '??????', '??????',
                         '??????', '??????', '??????', '?????????', '??????', '???', '??????', '??????', '???', '??????',
                         '???', '???', '???', '??????', '???', '???', '??????', '???', '??????', '???', '??????', '???', '??????',
                         '???', '??????', '??????', '??????', '???', '??????', '???', '???']
    adverb_dict['more'] = [ '?????????', '???', '???', '??????', '????????????', '??????', '???', '??????', '???', '??????',
                          '??????', '?????????', '??????', '??????', '??????', '???', '??????', '???', '??????', '??????',
                          '???', '???', '??????', '??????', '?????????', '??????', '??????', '??????', '??????', '?????????',
                          '??????', '??????', '??????', '???', '??????']
    adverb_dict['ish'] = ['????????????', '????????????', '???', '??????', '???', '????????????', '???', '??????', '??????', '??????', '??????',
                        '???', '???', '??????', '??????', '??????', '??????', '???', '??????', '??????', '???', '??????', '??????',
                        '??????', '?????????', '??????', '??????', '?????????', '??????']
    adverb_dict['insufficiently'] = ['??????', '??????', '????????????', '??????', '?????????',
                                   '???', '?????????', '??????', '???', '??????', '???', '??????']
    return adverb_dict

def get_privative_list(negitive_words_dir):
    """
    get privative_list
    :return: [...]
    """
    file = open(negitive_words_dir, 'r', encoding='utf-8').read()
    stop_privative_list = file.split('\n')
    return stop_privative_list

def load_data(data_dir):
    """
    get source weibo data
    :return: {'uuid':['text', '0 or 1 or 2']}
    """
    wb = load_workbook(data_dir)
    ws = wb.get_sheet_by_name('Sheet1')
    data_dict = {}
    for i in range(1, ws.max_row):
        try:
            weibo_score = [ws.cell(row = i+1, column = 4).value, ws.cell(row = i+1, column = 9).value]
            data_dict[ws.cell(row = i+1, column = 1).value] = weibo_score
        except KeyError:
            print ("KeyError_Line:: ", i+1)
            continue
    return data_dict

def Boson_dict(boson_dict_dir):
    """
    get Boson_Emotional_dict
    :return: {'Words':'Boson_Score'}
    """
    Boson_file = open(boson_dict_dir, 'r',
                encoding='utf-8').read()
    Boson_dict = {}
    Boson_file = Boson_file.split('\n')
    for boson in Boson_file:
        couple = boson.split(' ')
        try:
            Boson_dict[couple[0]] = round(float(couple[1]),4)
        except IndexError:
            continue
    return Boson_dict

def get_stop_word(stop_word_dir):
    """
    get Stop_word_list
    :return: [...]
    """
    file = open(stop_word_dir, 'r', encoding='utf-8').read()
    stop_word_list = file.split('\n')
    return stop_word_list

def get_Boson_score(weibo, Boson_dict, stop_word_list, compile_sentiment_dict, sentiment_dict, adverb_dict, privative_list):
    """
    get Boson_Model_Emotion_Score
    Modify by Emotional_Dictionary,but where is it?
    :param weibo:{'uuid':['text', '0 or 1 or 2']}
    :param Boson_dict:{'Words':'Boson_Score'}
    :param stop_word_list:[...]
    :param compile_sentiment_dict:{'compile phrase':score}
    :param sentiment_dict:{'pos':[...],'neg':[...]}
    :param adverb_dict:{'extreme':[...],...,'insufficently':[...]}
    :param privative_list:[...]
    :return:float(score)
    """
    seg_list = jieba.cut(weibo, cut_all = False)
    seg_to_stop_list = [seg for seg in seg_list if seg not in stop_word_list]
    re_privative_list = [seg for seg in seg_to_stop_list if seg in privative_list]
    re_compile_sentiment_list = [[seg,compile_sentiment_dict[seg]] for seg in seg_to_stop_list if seg in compile_sentiment_dict.keys()]
    re_adverb_list = [seg for seg in seg_to_stop_list for key in adverb_dict if seg in adverb_dict[key]]
    Emotional_Score = 0
    weight = 1
    for word in seg_to_stop_list:
        if word in re_privative_list:
            weight = -1*weight
        elif word in re_compile_sentiment_list:
            Emotional_Score += re_compile_sentiment_list[word][1]
        elif word in re_adverb_list:
            if word in adverb_dict['extreme']:
                Emotional_Score += 2*Emotional_Score
            if word in adverb_dict['very']:
                Emotional_Score += 1.75*Emotional_Score
            if word in adverb_dict['more']:
                Emotional_Score += 1.5*Emotional_Score
            if word in adverb_dict['ish']:
                Emotional_Score += 1.3*Emotional_Score
            if word in adverb_dict['insufficiently']:
                Emotional_Score += 1.1*Emotional_Score
        else:
            for key in sentiment_dict:
                if word in sentiment_dict[key]:
                    if word in Boson_dict:
                        if Boson_dict[word] < 0 and key == 'neg':
                            Emotional_Score += Boson_dict[word]
                        elif Boson_dict[word] > 0 and key == 'pos':
                            Emotional_Score += Boson_dict[word]
                        else:
                            Emotional_Score += -1*Boson_dict[word]
    # try:
    #     Emotional_Score = round(Emotional_Score/no_neutral, 4)
    # except ZeroDivisionError:
    #     Emotional_Score = 0
    return Emotional_Score

def Accuracy_Boson_Model(data_dict, res_dict):
    """
    Accuracy_Boson_Model
    :param data_dict:{'uuid':['text', '0 or 1 or 2']}
    :param res_dict:{uuid:score,...}
    :return: several accuracy score
    """
    res = 0
    neutral = 0
    res_neu = 0
    pos = 0
    res_pos = 0
    neg = 0
    res_neg = 0
    for uuid in data_dict:
        res_score = res_dict[uuid]
        org_score = int(data_dict[uuid][1])
        if org_score == 1:
            neutral += 1
            if res_score == org_score:
                res_neu += 1
        if org_score == 0:
            neg += 1
            if res_score == org_score:
                res_neg += 1
        if org_score == 2:
            pos += 1
            if res_score == org_score:
                res_pos += 1
        if res_score == org_score:
            res += 1
    ratio_total= res/len(data_dict)
    ratio_pos = res_pos/pos
    ratio_neu = res_neu/neutral
    ratio_neg = res_neg/neg
    # print ('res_neg, neg:', res_neg, neg)
    return ratio_total, ratio_pos, ratio_neu, ratio_neg

def boson_Dict_Model(data_dict, Boson_dict, compile_sentiment_dict, sentiment_dict, adverb_dict, stop_word_dir, negitive_words_dir):
    """
    Boson_Model Adjust factors
    :param data_dict: {'uuid':['text', '0 or 1 or 2']}
    :param Boson_dict: {'Words':'Boson_Score'}
    :param compile_sentiment_dict: {'compile phrase':score}
    :param sentiment_dict: {'pos':[...],'neg':[...]}
    :param adverb_dict: {'extreme':[...],...,'insufficently':[...]}
    :return: return Boson method result
    """
    Boson_Model_Result_dict = {}
    stop_word_list = get_stop_word(stop_word_dir)
    privative_list = get_privative_list(negitive_words_dir)
    for uuid in data_dict:
        Boson_Model_Result_dict[uuid] = get_Boson_score(data_dict[uuid][0], Boson_dict, stop_word_list,compile_sentiment_dict, sentiment_dict,adverb_dict,privative_list)
    for uuid in Boson_Model_Result_dict:
        score = float(Boson_Model_Result_dict[uuid])
        # print (score)
        if score < lower:
            Boson_Model_Result_dict[uuid] = 0
        if score > upper:
            Boson_Model_Result_dict[uuid] = 2
        if score >= lower and score <= upper:
            Boson_Model_Result_dict[uuid] = 1
    Ratio, pos_ratio, neu_ratio, neg_ratio = Accuracy_Boson_Model(data_dict, Boson_Model_Result_dict)
    return Ratio, pos_ratio, neu_ratio, neg_ratio,Boson_Model_Result_dict

def Boson_Dict_Method(data_dir, boson_dir, sentiment_dir1, sentiment_dir2, sentiment_dir3,
                      sentiment_dir4, stop_word_dir, negitive_words_dir):
    """
    :return: several accuracy ratio score
    """
    start = time.time()
    data_dict = load_data(data_dir)
    # Boson_dict = {'Words':'Boson_Score'}
    boson_dict = Boson_dict(boson_dir)
    compile_sentiment_dict, sentiment_dict = get_sentiment_dict(boson_dict, sentiment_dir1, sentiment_dir2, sentiment_dir3, sentiment_dir4)
    adverb_dict = get_adverb_dict()
    # add into jieba dict
    load_jieba(compile_sentiment_dict)
    ratio, ratio_pos, ratio_neu, ratio_neg, Boson_Model_Result_dict = boson_Dict_Model(data_dict, boson_dict,
                        compile_sentiment_dict, sentiment_dict, adverb_dict, stop_word_dir, negitive_words_dir)
    print('Bonson_Model_Accuracy: ', ratio)
    print('Pos_Recall: ', ratio_pos)
    print('Neu_Recall: ', ratio_neu)
    print('Neg_Model_Recall: ', ratio_neg)
    end = time.time()
    print("Total Time: ", end - start)
    return ratio

if __name__ == "__main__":
    globals()
    # select the best config
    upper_list = [0.5, 1, 1.5, 2, 2.5, 3, 3.5, 4]
    lower_list = [-0.5, -1, -1.5, -2, -2.5, -3, -3.5, -4]
    ratio_list = []
    for i in tqdm(range(len(upper_list))):
        upper = upper_list[i]
        for j in range(len(upper_list)):
            lower = lower_list[j]
            ratio = Boson_Dict_Method(bc.data, bc.boson_dict_dir, bc.sentiment_dir1,
                bc.sentiment_dir2, bc.sentiment_dir3, bc.sentiment_dir4, bc.stop_word_dir,bc.negitive_words_dir)
            ratio_list.append(ratio)

    # #????????????upper?????????lower???accuracy?????????

    # fig = plt.figure(figsize = (16,4))
    # plt.subplot(181)
    # x = lower_list
    # y = ratio_list[:8]
    # plt.plot(x,y,color = 'r', linestyle = '-')
    # plt.xlabel(u'lower parameters')
    # plt.ylabel(u'accuracy')
    # plt.title("0.5 upper")
    #
    # plt.subplot(182)
    # x = lower_list
    # y = ratio_list[8:16]
    # plt.plot(x, y, color='r', linestyle='-')
    # plt.xlabel(u'lower parameters')
    # plt.ylabel(u'accuracy')
    # plt.title("1.0 upper")
    #
    # plt.subplot(183)
    # x = lower_list
    # y = ratio_list[16:24]
    # plt.plot(x, y, color='r', linestyle='-')
    # plt.xlabel(u'lower parameters')
    # plt.ylabel(u'accuracy')
    # plt.title("1.5 upper")
    #
    # plt.subplot(184)
    # x = lower_list
    # y = ratio_list[24:32]
    # plt.plot(x, y, color='r', linestyle='-')
    # plt.xlabel(u'lower parameters')
    # plt.ylabel(u'accuracy')
    # plt.title("2.0 upper")
    #
    # plt.subplot(185)
    # x = lower_list
    # y = ratio_list[32:40]
    # plt.plot(x, y, color='r', linestyle='-')
    # plt.xlabel(u'lower parameters')
    # plt.ylabel(u'accuracy')
    # plt.title("2.5 upper")
    #
    # plt.subplot(186)
    # x = lower_list
    # y = ratio_list[40:48]
    # plt.plot(x, y, color='r', linestyle='-')
    # plt.xlabel(u'lower parameters')
    # plt.ylabel(u'accuracy')
    # plt.title("3.0 upper")
    #
    # plt.subplot(187)
    # x = lower_list
    # y = ratio_list[48:56]
    # plt.plot(x, y, color='r', linestyle='-')
    # plt.xlabel(u'lower parameters')
    # plt.ylabel(u'accuracy')
    # plt.title("3.5 upper")
    #
    # plt.subplot(188)
    # x = lower_list
    # y = ratio_list[56:64]
    # plt.plot(x, y, color='r', linestyle='-')
    # plt.xlabel(u'lower parameters')
    # plt.ylabel(u'accuracy')
    # plt.title("4.0 upper")
    # plt.show()




