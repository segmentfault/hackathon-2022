# -*- coding: utf-8 -*-
'''
This code is implemented by Chauby, it is free for everyone.
Email: chaubyZou@163.com
'''

#%% import package
import numpy as np
from scipy.interpolate import interp1d
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
#补充
from scipy.interpolate import make_interp_spline
from openpyxl import load_workbook

from cs import CanonicalSystem
# coding=UTF-8
import pandas as pd


#%% define discrete dmp
class dmp_discrete():
    def __init__(self, n_dmps=1, n_bfs=100, dt=0, alpha_y=None, beta_y=None, **kwargs):
        self.n_dmps = n_dmps # number of data dimensions, one dmp for one degree
        self.n_bfs = n_bfs # number of basis functions
        self.dt = dt

        self.y0 = np.zeros(n_dmps)  # for multiple dimensions
        self.goal = np.ones(n_dmps) # for multiple dimensions

        self.alpha_y = np.ones(n_dmps) * 60.0 if alpha_y is None else alpha_y
        self.beta_y = self.alpha_y / 4.0 if beta_y is None else beta_y
        self.tau = 1.0

        self.w = np.zeros((n_dmps, n_bfs)) # weights for forcing term
        self.psi_centers = np.zeros(self.n_bfs) # centers over canonical system for Gaussian basis functions
        self.psi_h = np.zeros(self.n_bfs) # variance over canonical system for Gaussian basis functions

        # canonical system
        self.cs = CanonicalSystem(dt=self.dt, **kwargs)
        self.timesteps = round(self.cs.run_time / self.dt)

        # generate centers for Gaussian basis functions
        self.generate_centers()

        # self.h = np.ones(self.n_bfs) * self.n_bfs / self.psi_centers # original
        self.h = np.ones(self.n_bfs) * self.n_bfs**1.5 / self.psi_centers / self.cs.alpha_x # chose from trail and error

        # reset state
        self.reset_state()

    # Reset the system state
    def reset_state(self):
        self.y = self.y0.copy()
        self.dy = np.zeros(self.n_dmps)
        self.ddy = np.zeros(self.n_dmps)
        self.cs.reset_state()

    def generate_centers(self):
        t_centers = np.linspace(0, self.cs.run_time, self.n_bfs) # centers over time

        cs = self.cs
        x_track = cs.run() # get all x over run time
        t_track = np.linspace(0, cs.run_time, cs.timesteps) # get all time ticks over run time

        for n in range(len(t_centers)):
            for i, t in enumerate(t_track):
                if abs(t_centers[n] - t) <= cs.dt: # find the x center corresponding to the time center
                    self.psi_centers[n] = x_track[i]
        
        return self.psi_centers

    def generate_psi(self, x):
        if isinstance(x, np.ndarray):
            x = x[:, None]

        self.psi = np.exp(-self.h * (x - self.psi_centers)**2)

        return self.psi
    
    def generate_weights(self, f_target):
        x_track = self.cs.run()
        psi_track = self.generate_psi(x_track)

        for d in range(self.n_dmps):
            # ------------ Original DMP in Schaal 2002
            # delta = self.goal[d] - self.y0[d]
            
            # ------------ Modified DMP in Schaal 2008
            delta = 1.0

            for b in range(self.n_bfs):
                # as both number and denom has x(g-y_0) term, thus we can simplify the calculation process
                numer = np.sum(x_track * psi_track[:,b] * f_target[:,d])
                denom = np.sum(x_track**2 * psi_track[:,b])
                # numer = np.sum(psi_track[:,b] * f_target[:,d]) # the simpler calculation
                # denom = np.sum(x_track * psi_track[:,b])
                self.w[d, b] = numer / (denom*delta)
        print(1)
        print(self.w)
        print(1)
        return self.w

    def learning(self, y_demo, plot=False):
        if y_demo.ndim == 1: # data is with only one dimension
            y_demo = y_demo.reshape(1, len(y_demo))

        self.y0 = y_demo[:,0].copy()
        self.goal = y_demo[:,-1].copy()
        self.y_demo = y_demo.copy()

        # interpolate the demonstrated trajectory to be the same length with timesteps
        x = np.linspace(0, self.cs.run_time, y_demo.shape[1])
        y = np.zeros((self.n_dmps, self.timesteps))
        for d in range(self.n_dmps):
            y_tmp = interp1d(x, y_demo[d])
            for t in range(self.timesteps):
                y[d, t] = y_tmp(t*self.dt)
        
        # calculate velocity and acceleration of y_demo

        # method 1: using gradient
        dy_demo = np.gradient(y, axis=1) / self.dt
        ddy_demo = np.gradient(dy_demo, axis=1) / self.dt

        # method 2: using diff
        # dy_demo = np.diff(y) / self.dt
        # # let the first gradient same as the second gradient
        # dy_demo = np.hstack((np.zeros((self.n_dmps, 1)), dy_demo)) # Not sure if is it a bug?
        # # dy_demo = np.hstack((dy_demo[:,0].reshape(self.n_dmps, 1), dy_demo))

        # ddy_demo = np.diff(dy_demo) / self.dt
        # # let the first gradient same as the second gradient
        # ddy_demo = np.hstack((np.zeros((self.n_dmps, 1)), ddy_demo))
        # # ddy_demo = np.hstack((ddy_demo[:,0].reshape(self.n_dmps, 1), ddy_demo))

        x_track = self.cs.run()
        f_target = np.zeros((y_demo.shape[1], self.n_dmps))
        for d in range(self.n_dmps):
            # ---------- Original DMP in Schaal 2002
            # f_target[:,d] = ddy_demo[d] - self.alpha_y[d]*(self.beta_y[d]*(self.goal[d] - y_demo[d]) - dy_demo[d])

            # ---------- Modified DMP in Schaal 2008, fixed the problem of g-y_0 -> 0
            k = self.alpha_y[d]
            f_target[:,d] = (ddy_demo[d] - self.alpha_y[d]*(self.beta_y[d]*(self.goal[d] - y_demo[d]) - dy_demo[d]))/k + x_track*(self.goal[d] - self.y0[d])
        
        self.generate_weights(f_target)

        if plot is True:
            # plot the basis function activations
            plt.figure()
            plt.subplot(211)
            psi_track = self.generate_psi(self.cs.run())
            plt.plot(psi_track)
            plt.title('basis functions')

            # plot the desired forcing function vs approx
            plt.subplot(212)
            plt.plot(f_target[:,0])
            plt.plot(np.sum(psi_track * self.w[0], axis=1) * self.dt)
            plt.legend(['f_target', 'w*psi'])
            plt.title('DMP forcing function')
            plt.tight_layout()
            plt.show()

        # reset state
        self.reset_state()


    def reproduce(self, tau=None, initial=None, goal=None):
        # set temporal scaling
        if tau == None:
            timesteps = self.timesteps
        else:
            timesteps = round(self.timesteps/tau)

        # set initial state
        if initial != None:
            self.y0 = initial
        
        # set goal state
        if goal != None:
            self.goal = goal
        
        # reset state
        self.reset_state()

        y_reproduce = np.zeros((timesteps, self.n_dmps))
        dy_reproduce = np.zeros((timesteps, self.n_dmps))
        ddy_reproduce = np.zeros((timesteps, self.n_dmps))

        for t in range(timesteps):
            y_reproduce[t], dy_reproduce[t], ddy_reproduce[t] = self.step(tau=tau)
        
        return y_reproduce, dy_reproduce, ddy_reproduce

    def step(self, tau=None):
        # run canonical system
        if tau == None:
            tau = self.tau
        x = self.cs.step_discrete(tau)

        # generate basis function activation
        psi = self.generate_psi(x)

        for d in range(self.n_dmps):
            # generate forcing term
            # ------------ Original DMP in Schaal 2002
            # f = np.dot(psi, self.w[d])*x*(self.goal[d] - self.y0[d]) / np.sum(psi)

            # ---------- Modified DMP in Schaal 2008, fixed the problem of g-y_0 -> 0
            k = self.alpha_y[d]
            f = k*(np.dot(psi, self.w[d])*x / np.sum(psi)) - k*(self.goal[d] - self.y0[d])*x

            # generate reproduced trajectory
            self.ddy[d] = self.alpha_y[d]*(self.beta_y[d]*(self.goal[d] - self.y[d]) - self.dy[d]) + f
            self.dy[d] += tau*self.ddy[d]*self.dt
            self.y[d] += tau*self.dy[d]*self.dt
        
        return self.y, self.dy, self.ddy


#%% test code
if __name__ == "__main__":
    
    book = load_workbook(filename=r"C:/Users/admin/Desktop/天宇师兄/test002.xlsx")
    #读取名字为Sheet1的表
    sheet1 = book.get_sheet_by_name("trajectory")
    #用于存储数据的数组
    data1= []
    row_num = 1
    while row_num <= 500 :
    #将表中第一列的1-100行数据写入data数组中
        data1.append(sheet1.cell(row=row_num, column=7).value)
        row_num = row_num + 1

    sheet2 = book.get_sheet_by_name("trajectory")
    #用于存储数据的数组
    data2= []
    row_num = 1
    while row_num <= 500 :
    #将表中第一列的1-100行数据写入data数组中
        data2.append(sheet2.cell(row=row_num, column=8).value)
        row_num = row_num + 1
    sheet3 = book.get_sheet_by_name("trajectory")
    #用于存储数据的数组
    data3= []
    row_num = 1
    while row_num <= 500 :
    #将表中第一列的1-100行数据写入data数组中
        data3.append(sheet3.cell(row=row_num, column=9).value)
        row_num = row_num + 1


    data_len = 500
    y_demo = np.zeros((3, data_len))
    # print(y_demo)
    y_demo[0,:] = np.array(data1)
    y_demo[1,:] = np.array(data2)
    y_demo[2,:] = np.array(data3)
   
    
    # DMP learning
    dmp = dmp_discrete(n_dmps=y_demo.shape[0], n_bfs=10000, dt=1.0/data_len)
    dmp.learning(y_demo, plot=False)

    # reproduce learned trajectory
    y_reproduce, dy_reproduce, ddy_reproduce = dmp.reproduce()

    plt.figure(figsize=(10, 5))
    plt.plot(y_demo[0,:], 'g', label='demo1')
    plt.plot(y_reproduce[:,0], 'r--', label='reproduce1')
    # plt.plot(y_reproduce_2[:,0], 'r-.', label='reproduce12')
    plt.plot(y_demo[1,:], 'b', label='demo2')
    plt.plot(y_reproduce[:,1], 'm--', label='reproduce2')
    # plt.plot(y_reproduce_2[:,1], 'm-.', label='reproduce22')
    plt.plot(y_demo[2,:], 'c', label='demo3')
    plt.plot(y_reproduce[:,2], 'y--', label='reproduce3')
    font1 = {'family' : 'Times New Roman',
    'weight' : 'normal',
    'size'   : 22,
    }
    plt.xticks(fontproperties = 'Times New Roman', size = 35)
    plt.yticks(fontproperties = 'Times New Roman', size = 35)
    plt.legend(prop=font1)
    plt.grid()
    plt.xlabel('Timesteps', fontproperties = 'Times New Roman',fontsize=35)
    plt.ylabel('Position', fontproperties = 'Times New Roman',fontsize=35)
    plt.show()
    fig=plt.figure()
    ax1 = Axes3D(fig)
    ax1.plot3D(y_demo[0,:],y_demo[1,:],y_demo[2,:],'y', label='demoXYZ')    #绘制空间曲线
    ax1.plot3D(y_reproduce[:,0],y_reproduce[:,1],y_reproduce[:,2],'k--', label='reproduceXYZ') 

    font1 = {'family' : 'Times New Roman',
    'weight' : 'normal',
    'size'   : 22,
    }
    ax1.legend(prop=font1)
    plt.grid()
    ax1.set_xlabel('X position', fontproperties = 'Times New Roman',fontsize=35)
    ax1.set_ylabel('Y position', fontproperties = 'Times New Roman',fontsize=35)
    ax1.set_zlabel('Z position', fontproperties = 'Times New Roman',fontsize=35)
    plt.show()

    plt.figure(figsize=(10, 5))
    
    plt.plot(dy_reproduce[:,0], 'r--', label='reproduce1')
    # plt.plot(y_reproduce_2[:,0], 'r-.', label='reproduce12')
    
    plt.plot(dy_reproduce[:,1], 'm--', label='reproduce2')
    # plt.plot(y_reproduce_2[:,1], 'm-.', label='reproduce22')
    
    plt.plot(dy_reproduce[:,2], 'y--', label='reproduce3')
    font1 = {'family' : 'Times New Roman',
    'weight' : 'normal',
    'size'   : 22,
    }
    plt.xticks(fontproperties = 'Times New Roman', size = 35)
    plt.yticks(fontproperties = 'Times New Roman', size = 35)
    plt.legend(prop=font1)
    plt.grid()
    plt.xlabel('Timesteps', fontproperties = 'Times New Roman',fontsize=35)
    plt.ylabel('Velocity', fontproperties = 'Times New Roman',fontsize=35)
    plt.show()




